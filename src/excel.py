import datetime
from pathlib import Path

import xlsxwriter
from openpyxl import load_workbook


def open_excel(name_file: str) -> list[list] | None:
    """функция получения сырого массива из excel файла"""
    try:
        wb = load_workbook(filename=name_file, read_only=True)
    except Exception as ex:
        print(f'ошибка открытия файла: {name_file} \n {ex}')
        return None
    try:
        list_name = 'Лист1'
        ws = wb[list_name]
    except Exception as ex:
        print(f'ошибка доступа к листу: {list_name} \n {ex}')
        return None
    # массив для данных из excel
    mass_excel = []
    for row in ws.rows:
        # добавить строку в массив
        mass_excel.append([cell.value for cell in row])
    # закрытие документа
    wb.close()

    # вернуть массив
    return mass_excel


def f_save_xlsx(name_file: str, dir: str, header: list, data: list):
    """функция записи массива в excel файла"""

    data.insert(0, header)

    now = datetime.datetime.now()
    name_f = f'{name_file} ' + str(now).replace(':', '_') + '.xlsx'

    path_for_file = dir

    if Path(path_for_file).exists() is False:
        path_dir = Path(path_for_file)
        path_dir.mkdir()

    name_file = Path(path_for_file, name_f)

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(name_file)
    worksheet = workbook.add_worksheet()

    # Start from the first cell. Rows and columns are zero indexed.

    # Iterate over the data and write it out row by row.
    for count_row, line_r in enumerate(data):
        for count_col, line_c in enumerate(line_r):
            worksheet.write(count_row, count_col, line_c)

    workbook.close()

    print(f'save file comlite: {name_f}')
